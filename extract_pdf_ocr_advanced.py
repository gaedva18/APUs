#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script avanzado con OCR para extraer datos de PDFs escaneados
Requiere: pip install pdfplumber pytesseract pdf2image openpyxl pandas pillow
"""

import pdfplumber
import pandas as pd
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

class AdvancedAPUExtractor:
    """Extrae datos de APUs desde PDFs con OCR avanzado"""
    
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.data = []
        self.patterns = self._compile_patterns()
        
    def _compile_patterns(self):
        """Compila patrones regex para APUs"""
        return {
            'codigo': r'(?:APU|RUBRO|CÓDIGO)[\s:]*([A-Z0-9\-\.]+)',
            'descripcion': r'(?:DESCRIPCIÓN|DESC)[\s:]*([^\n]+)',
            'cantidad': r'(?:CANTIDAD|CANT)[\s:]*(\d+\.?\d*)',
            'unitario': r'(?:UNITARIO|VR\.?\s*UNI)[\s:]*\$?(\d+\.?\d*)',
            'total': r'(?:TOTAL|SUBTOTAL)[\s:]*\$?(\d+\.?\d*)',
            'apu_line': r'^([A-Z0-9\-\.]+)\s+(.+?)\s+(\d+\.?\d*)\s+\$?(\d+\.?\d*)\s+\$?(\d+\.?\d*)$'
        }
    
    def extract(self):
        """Extrae datos usando múltiples estrategias"""
        print(f"📄 Procesando: {self.pdf_path}")
        
        # Intenta primero con pdfplumber
        self._extract_with_pdfplumber()
        
        # Si no hay datos, intenta con OCR
        if not self.data:
            print("   ⚠️  Usando OCR para PDFs escaneados...")
            self._extract_with_ocr()
        
        print(f"   ✓ Registros extraídos: {len(self.data)}")
        return self.data
    
    def _extract_with_pdfplumber(self):
        """Extrae datos con pdfplumber"""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    # Intenta extraer tablas
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            self._process_table_data(table, page_num)
                    
                    # Extrae texto
                    text = page.extract_text()
                    if text and not tables:
                        self._extract_from_text(text, page_num)
                    
        except Exception as e:
            print(f"   ⚠️  Error con pdfplumber: {e}")
    
    def _extract_with_ocr(self):
        """Extrae datos usando OCR (Tesseract)"""
        try:
            images = convert_from_path(self.pdf_path, dpi=300)
            
            for page_num, image in enumerate(images, 1):
                # Aplica OCR
                text = pytesseract.image_to_string(image, lang='spa+eng')
                self._extract_from_text(text, page_num)
                
        except Exception as e:
            print(f"   ⚠️  Error con OCR: {e}")
            print("   💡 Instala Tesseract: apt-get install tesseract-ocr")
    
    def _process_table_data(self, table, page_num):
        """Procesa datos de tabla"""
        if len(table) < 2:
            return
        
        headers = table[0]
        for row in table[1:]:
            if any(row):
                record = {}
                record['Página'] = page_num
                record['Tipo'] = 'Tabla'
                
                for i, header in enumerate(headers):
                    key = str(header).strip() if header else f'Col_{i}'
                    value = str(row[i]).strip() if i < len(row) and row[i] else ''
                    record[key] = value
                
                if len(record) > 2:
                    self.data.append(record)
    
    def _extract_from_text(self, text, page_num):
        """Extrae datos estructurados del texto"""
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or len(line) < 3:
                continue
            
            # Intenta extraer como línea de APU
            match = re.search(self.patterns['apu_line'], line)
            if match:
                record = {
                    'Página': page_num,
                    'Tipo': 'APU',
                    'Código': match.group(1),
                    'Descripción': match.group(2),
                    'Cantidad': match.group(3),
                    'Unitario': match.group(4),
                    'Total': match.group(5)
                }
                self.data.append(record)
            else:
                # Busca patrones individuales
                record = {'Página': page_num, 'Tipo': 'Texto'}
                
                for pattern_name, pattern in self.patterns.items():
                    if pattern_name != 'apu_line':
                        match = re.search(pattern, line)
                        if match:
                            record[pattern_name.capitalize()] = match.group(1)
                
                if len(record) > 2:
                    record['Contenido'] = line
                    self.data.append(record)


class DataProcessor:
    """Procesa y limpia datos extraídos"""
    
    @staticmethod
    def create_dataframe(data):
        """Crea DataFrame desde datos extraídos"""
        if not data:
            return pd.DataFrame()
        
        df = pd.DataFrame(data)
        
        # Limpia datos numéricos
        numeric_columns = ['Cantidad', 'Unitario', 'Total', 'cantidad', 'unitario', 'total']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace('$', '').str.replace(',', '.'),
                    errors='coerce'
                )
        
        return df
    
    @staticmethod
    def consolidate_data(dataframes_dict):
        """Consolida múltiples DataFrames"""
        all_data = []
        
        for source, df in dataframes_dict.items():
            if not df.empty:
                df['Fuente'] = source
                all_data.append(df)
        
        if all_data:
            return pd.concat(all_data, ignore_index=True)
        return pd.DataFrame()


class ExcelExporterAdvanced:
    """Exporta datos a Excel con formato avanzado"""
    
    def __init__(self, output_path):
        self.output_path = output_path
    
    def export(self, data_frames_dict):
        """Exporta múltiples DataFrames"""
        try:
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Exporta hojas individuales
                for sheet_name, df in data_frames_dict.items():
                    if not df.empty:
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        worksheet = writer.sheets[sheet_name[:31]]
                        self._format_worksheet(worksheet, df)
                
                # Crea hoja consolidada si hay múltiples fuentes
                if len(data_frames_dict) > 1:
                    consolidated = self._consolidate(data_frames_dict)
                    if not consolidated.empty:
                        consolidated.to_excel(writer, sheet_name='Consolidado', index=False)
                        self._format_worksheet(writer.sheets['Consolidado'], consolidated)
            
            print(f"✅ Archivo exportado: {self.output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error al exportar: {e}")
            return False
    
    def _consolidate(self, dataframes_dict):
        """Consolida datos de múltiples fuentes"""
        processor = DataProcessor()
        return processor.consolidate_data(dataframes_dict)
    
    def _format_worksheet(self, worksheet, df):
        """Aplica formato profesional"""
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Datos
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Formato de números
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
        
        # Ancho automático
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Función principal"""
    print("=" * 70)
    print("EXTRACTOR AVANZADO DE APUs - PDF a Excel con OCR")
    print("=" * 70)
    
    pdf_files = ['003_opt.pdf', '004.pdf']
    output_file = 'APUs_Extraidos_Avanzado.xlsx'
    
    all_dataframes = {}
    
    # Extrae de cada PDF
    for pdf_file in pdf_files:
        try:
            extractor = AdvancedAPUExtractor(pdf_file)
            data = extractor.extract()
            
            if data:
                processor = DataProcessor()
                df = processor.create_dataframe(data)
                
                sheet_name = pdf_file.replace('.pdf', '').replace('_', ' ')
                all_dataframes[sheet_name] = df
                
                print(f"\n📊 Resumen: {sheet_name}")
                print(f"   Registros: {len(df)}")
                print(f"   Columnas: {', '.join(df.columns.tolist())}")
                if not df.empty:
                    print("\n   Primeras filas:")
                    print(df.head(3).to_string())
                
        except FileNotFoundError:
            print(f"⚠️  Archivo no encontrado: {pdf_file}")
        except Exception as e:
            print(f"⚠️  Error procesando {pdf_file}: {e}")
    
    # Exporta a Excel
    if all_dataframes:
        exporter = ExcelExporterAdvanced(output_file)
        success = exporter.export(all_dataframes)
        
        if success:
            print("\n" + "=" * 70)
            print("✅ PROCESO COMPLETADO EXITOSAMENTE")
            print(f"📁 Archivo guardado: {output_file}")
            print("=" * 70)
            return 0
    else:
        print("\n❌ No se extrajeron datos de los PDFs")
        return 1
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
