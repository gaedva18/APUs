#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para extraer datos de PDFs de APUs y exportar a Excel
Requiere: pip install pdfplumber openpyxl pandas
"""

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import sys

class APUExtractor:
    """Extrae datos de APUs desde PDFs"""
    
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.data = []
        
    def extract(self):
        """Extrae tablas del PDF"""
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                print(f"📄 Procesando: {self.pdf_path}")
                print(f"   Total de páginas: {len(pdf.pages)}")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    tables = page.extract_tables()
                    if tables:
                        print(f"   ✓ Página {page_num}: {len(tables)} tabla(s) encontrada(s)")
                        for table in tables:
                            self._process_table(table, page_num)
                    else:
                        # Intenta extraer texto si no hay tablas
                        text = page.extract_text()
                        if text:
                            self._process_text(text, page_num)
                
                print(f"\n✅ Total de registros extraídos: {len(self.data)}")
                return self.data
                
        except Exception as e:
            print(f"❌ Error al procesar PDF: {e}")
            return []
    
    def _process_table(self, table, page_num):
        """Procesa una tabla extraída"""
        if not table or len(table) < 2:
            return
            
        headers = table[0]
        for row in table[1:]:
            if any(row):  # Si la fila no está vacía
                record = self._create_record(headers, row, page_num)
                if record:
                    self.data.append(record)
    
    def _process_text(self, text, page_num):
        """Procesa texto cuando no hay tablas"""
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if line and len(line) > 5:
                # Busca patrones numéricos o de APU
                if re.search(r'\d+', line):
                    self.data.append({
                        'Página': page_num,
                        'Contenido': line,
                        'Tipo': 'Texto extraído'
                    })
    
    def _create_record(self, headers, row, page_num):
        """Crea un registro desde encabezados y fila"""
        record = {'Página': page_num}
        
        for i, header in enumerate(headers):
            if i < len(row):
                header_clean = str(header).strip() if header else f'Columna_{i}'
                value = str(row[i]).strip() if row[i] else ''
                record[header_clean] = value
        
        return record if len(record) > 1 else None


class ExcelExporter:
    """Exporta datos a Excel con formato"""
    
    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        
    def export(self, data_frames_dict):
        """
        Exporta múltiples DataFrames a hojas de Excel
        
        Args:
            data_frames_dict: Dict con {nombre_hoja: DataFrame}
        """
        try:
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                for sheet_name, df in data_frames_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Aplica formato a la hoja
                    worksheet = writer.sheets[sheet_name]
                    self._format_worksheet(worksheet, df)
            
            print(f"✅ Archivo exportado: {self.output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error al exportar: {e}")
            return False
    
    def _format_worksheet(self, worksheet, df):
        """Aplica formato a la hoja"""
        # Estilo de encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatea encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Formatea datos
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Ajusta ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Función principal"""
    print("=" * 60)
    print("EXTRACTOR DE APUs - PDF a Excel")
    print("=" * 60)
    
    # Archivos
    pdf_files = ['003_opt.pdf', '004.pdf']
    output_file = 'APUs_Extraidos.xlsx'
    
    all_data = {}
    
    # Extrae de cada PDF
    for pdf_file in pdf_files:
        try:
            extractor = APUExtractor(pdf_file)
            data = extractor.extract()
            
            if data:
                sheet_name = pdf_file.replace('.pdf', '').replace('_', ' ')[:31]  # Max 31 chars
                df = pd.DataFrame(data)
                all_data[sheet_name] = df
                
                print(f"\n📊 Resumen de '{sheet_name}':")
                print(df.head())
                
        except FileNotFoundError:
            print(f"⚠️  Archivo no encontrado: {pdf_file}")
    
    # Exporta a Excel
    if all_data:
        exporter = ExcelExporter('formato.xlsx', output_file)
        success = exporter.export(all_data)
        
        if success:
            print("\n" + "=" * 60)
            print("✅ PROCESO COMPLETADO EXITOSAMENTE")
            print(f"📁 Archivo guardado: {output_file}")
            print("=" * 60)
    else:
        print("\n❌ No se extrajeron datos de los PDFs")
        sys.exit(1)


if __name__ == '__main__':
    main()
